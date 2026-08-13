import io
import re
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.services.scoring_service import PROFILE_SCORE_VERSION


EXCEL_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument."
    "spreadsheetml.sheet"
)

CANDIDATE_HEADERS = (
    "Rank",
    "Candidate ID",
    "Candidate Name",
    "Candidate Level",
    "Score Version",
    "Candidate Profile Score",
    "Legacy Technical Profile Score",
    "Profile Rule Score",
    "Legacy Technical Rule Score",
    "AI Analysis Score",
    "Pipeline Stage",
    "AI Status",
    "Summary",
    "Created At",
)

PROFILE_BREAKDOWN_HEADERS = (
    "Candidate ID",
    "Candidate Name",
    "Professional Experience",
    "Achievements & Impact",
    "Competencies & Domain Expertise",
    "Certifications & Credentials",
    "Education",
    "Leadership & Responsibility",
    "Evidence Quality",
    "Profile Rule Score",
)

LEGACY_BREAKDOWN_HEADERS = (
    "Candidate ID",
    "Candidate Name",
    "Python",
    "SQL",
    "Backend",
    "DevOps",
    "AI Domain",
    "Data Domain",
    "Backend Domain",
    "Experience",
    "Projects",
    "Engineering Signal",
    "Legacy Technical Profile Score",
    "Legacy Technical Rule Score",
)

TITLE_FILL = "17365D"
HEADER_FILL = "2F5597"
HEADER_TEXT = "FFFFFF"
BAND_FILL = "F3F6FA"
BORDER_COLOR = "D9E2F3"
ILLEGAL_EXCEL_CHARACTERS = re.compile(
    r"[\x00-\x08\x0B\x0C\x0E-\x1F]"
)
EXCEL_CELL_TEXT_LIMIT = 32767


def build_candidates_workbook(
    candidates,
    generated_at: datetime | None = None,
) -> io.BytesIO:
    generated_at = generated_at or datetime.now(timezone.utc)
    candidate_list = list(candidates)

    workbook = Workbook()
    candidates_sheet = workbook.active
    candidates_sheet.title = "Candidates"
    profile_sheet = workbook.create_sheet("Score Breakdown")
    legacy_sheet = workbook.create_sheet("Legacy Scores")

    workbook.properties.creator = "ATS Resume Intelligence"
    workbook.properties.title = "Candidate Export Report"
    workbook.properties.subject = "Recruiter candidate export"

    _populate_candidates_sheet(
        candidates_sheet,
        candidate_list,
        generated_at,
    )
    _populate_profile_breakdown_sheet(
        profile_sheet,
        candidate_list,
        generated_at,
    )
    _populate_legacy_breakdown_sheet(
        legacy_sheet,
        candidate_list,
        generated_at,
    )

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _populate_candidates_sheet(sheet, candidates, generated_at):
    _add_report_metadata(
        sheet,
        len(CANDIDATE_HEADERS),
        generated_at,
        len(candidates),
    )
    _add_headers(sheet, CANDIDATE_HEADERS)

    for rank, candidate in enumerate(candidates, start=1):
        profile_v2 = _is_profile_v2(candidate)
        sheet.append([
            rank,
            candidate.id,
            _safe_text(candidate.name),
            _safe_text(candidate.candidate_level),
            PROFILE_SCORE_VERSION if profile_v2 else "technical_v1 (legacy)",
            _numeric(candidate.skill_score) if profile_v2 else None,
            None if profile_v2 else _numeric(candidate.skill_score),
            _numeric(candidate.rule_score) if profile_v2 else None,
            None if profile_v2 else _numeric(candidate.rule_score),
            _numeric(candidate.ai_score),
            _safe_text(candidate.candidate_stage),
            _safe_text(candidate.ai_status),
            _safe_text(candidate.summary),
            _excel_datetime(candidate.created_at),
        ])

    _finish_sheet(
        sheet,
        len(CANDIDATE_HEADERS),
        (8, 13, 28, 20, 23, 23, 29, 20, 28, 20, 18, 16, 60, 21),
        score_columns=(6, 7, 8, 9, 10),
        conditional_columns=(6, 7, 8, 9, 10),
        wrapped_columns=(3, 13),
        date_columns=(14,),
    )


def _populate_profile_breakdown_sheet(sheet, candidates, generated_at):
    profile_candidates = [
        candidate for candidate in candidates if _is_profile_v2(candidate)
    ]
    _add_report_metadata(
        sheet,
        len(PROFILE_BREAKDOWN_HEADERS),
        generated_at,
        len(profile_candidates),
    )
    _add_headers(sheet, PROFILE_BREAKDOWN_HEADERS)

    for candidate in profile_candidates:
        breakdown = candidate.score_breakdown or {}
        sheet.append([
            candidate.id,
            _safe_text(candidate.name),
            _breakdown_score(breakdown, "professional_experience"),
            _breakdown_score(breakdown, "achievements"),
            _breakdown_score(breakdown, "competencies"),
            _breakdown_score(breakdown, "certifications"),
            _breakdown_score(breakdown, "education"),
            _breakdown_score(breakdown, "leadership"),
            _breakdown_score(breakdown, "evidence_quality"),
            _numeric(candidate.rule_score),
        ])

    _finish_sheet(
        sheet,
        len(PROFILE_BREAKDOWN_HEADERS),
        (13, 28, 24, 22, 34, 29, 14, 30, 18, 20),
        score_columns=tuple(range(3, 11)),
        conditional_columns=(10,),
        wrapped_columns=(2,),
    )


def _populate_legacy_breakdown_sheet(sheet, candidates, generated_at):
    legacy_candidates = [
        candidate for candidate in candidates if not _is_profile_v2(candidate)
    ]
    _add_report_metadata(
        sheet,
        len(LEGACY_BREAKDOWN_HEADERS),
        generated_at,
        len(legacy_candidates),
    )
    _add_headers(sheet, LEGACY_BREAKDOWN_HEADERS)

    keys = (
        "python", "sql", "backend", "devops", "ai_domain",
        "data_domain", "backend_domain", "experience", "projects",
        "engineering_signal",
    )
    for candidate in legacy_candidates:
        breakdown = candidate.score_breakdown or {}
        sheet.append([
            candidate.id,
            _safe_text(candidate.name),
            *[_breakdown_score(breakdown, key) for key in keys],
            _numeric(candidate.skill_score),
            _numeric(candidate.rule_score),
        ])

    _finish_sheet(
        sheet,
        len(LEGACY_BREAKDOWN_HEADERS),
        (13, 28, 13, 13, 15, 13, 15, 16, 18, 16, 14, 22, 29, 28),
        score_columns=tuple(range(3, 15)),
        conditional_columns=(13, 14),
        wrapped_columns=(2,),
    )


def _add_report_metadata(sheet, column_count, generated_at, total):
    end_column = sheet.cell(row=1, column=column_count).column_letter
    for row in range(1, 5):
        sheet.merge_cells(f"A{row}:{end_column}{row}")

    generated_utc = generated_at
    if generated_at.tzinfo is not None:
        generated_utc = generated_at.astimezone(timezone.utc)

    sheet["A1"] = "ATS Resume Intelligence"
    sheet["A2"] = "Candidate Export Report"
    sheet["A3"] = "Generated: " + generated_utc.strftime(
        "%Y-%m-%d %H:%M UTC"
    )
    sheet["A4"] = f"Total Candidates: {total}"

    sheet["A1"].font = Font(
        name="Aptos Display", size=18, bold=True, color="FFFFFF"
    )
    sheet["A2"].font = Font(
        name="Aptos", size=12, bold=True, color="D9EAF7"
    )
    for row in (3, 4):
        sheet.cell(row=row, column=1).font = Font(
            name="Aptos", size=10, color="EAF2F8"
        )

    for row in range(1, 5):
        cell = sheet.cell(row=row, column=1)
        cell.fill = PatternFill("solid", fgColor=TITLE_FILL)
        cell.alignment = Alignment(vertical="center")

    sheet.row_dimensions[1].height = 28
    sheet.row_dimensions[2].height = 21
    sheet.row_dimensions[3].height = 18
    sheet.row_dimensions[4].height = 20
    sheet.row_dimensions[5].height = 8


def _add_headers(sheet, headers):
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=6, column=column, value=header)
        cell.font = Font(
            name="Aptos", size=10, bold=True, color=HEADER_TEXT
        )
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = Border(
            bottom=Side(style="medium", color=TITLE_FILL)
        )
    sheet.row_dimensions[6].height = 32


def _finish_sheet(
    sheet,
    column_count,
    widths,
    score_columns=(),
    conditional_columns=(),
    wrapped_columns=(),
    date_columns=(),
):
    last_row = max(sheet.max_row, 6)
    end_column = sheet.cell(row=6, column=column_count).column_letter
    sheet.freeze_panes = "A7"
    sheet.auto_filter.ref = f"A6:{end_column}{last_row}"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1

    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[
            sheet.cell(row=6, column=index).column_letter
        ].width = width

    light_side = Side(style="thin", color=BORDER_COLOR)
    for row in range(7, last_row + 1):
        if row % 2 == 0:
            for cell in sheet[row][:column_count]:
                cell.fill = PatternFill("solid", fgColor=BAND_FILL)
        for cell in sheet[row][:column_count]:
            cell.font = Font(name="Aptos", size=10, color="243447")
            cell.alignment = Alignment(vertical="top")
            cell.border = Border(bottom=light_side)
        sheet.row_dimensions[row].height = 36

    for column in score_columns:
        for row in range(7, last_row + 1):
            cell = sheet.cell(row=row, column=column)
            cell.number_format = "0"
            cell.alignment = Alignment(
                horizontal="center", vertical="center"
            )
    for column in conditional_columns:
        if last_row >= 7:
            column_letter = sheet.cell(row=6, column=column).column_letter
            sheet.conditional_formatting.add(
                f"{column_letter}7:{column_letter}{last_row}",
                ColorScaleRule(
                    start_type="num", start_value=0, start_color="F4CCCC",
                    mid_type="num", mid_value=60, mid_color="FFF2CC",
                    end_type="num", end_value=100, end_color="D9EAD3",
                ),
            )

    for column in wrapped_columns:
        for row in range(7, last_row + 1):
            sheet.cell(row=row, column=column).alignment = Alignment(
                vertical="top", wrap_text=True
            )

    for column in date_columns:
        for row in range(7, last_row + 1):
            cell = sheet.cell(row=row, column=column)
            cell.number_format = "yyyy-mm-dd hh:mm"
            cell.alignment = Alignment(
                horizontal="center", vertical="center"
            )


def _is_profile_v2(candidate) -> bool:
    breakdown = candidate.score_breakdown
    return (
        isinstance(breakdown, dict)
        and breakdown.get("score_version") == PROFILE_SCORE_VERSION
    )


def _safe_text(value) -> str:
    if value is None:
        return ""
    text = ILLEGAL_EXCEL_CHARACTERS.sub("", str(value))
    text = text[:EXCEL_CELL_TEXT_LIMIT]
    if text.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def _numeric(value):
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return None
    return value


def _breakdown_score(breakdown, key):
    if not isinstance(breakdown, dict):
        return None
    return _numeric(breakdown.get(key))


def _excel_datetime(value):
    if not isinstance(value, datetime):
        return None
    if value.tzinfo is not None:
        return value.astimezone(timezone.utc).replace(tzinfo=None)
    return value
