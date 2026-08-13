import os
from datetime import datetime, timezone
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import Mock

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

os.environ["TESTING"] = "true"

from main import app
from app.database.database import get_db
from app.services.excel_export_service import (
    CANDIDATE_HEADERS,
    EXCEL_CONTENT_TYPE,
    LEGACY_BREAKDOWN_HEADERS,
    PROFILE_BREAKDOWN_HEADERS,
    build_candidates_workbook,
)


@pytest.fixture
def client():

    with TestClient(app) as test_client:

        yield test_client


@pytest.fixture(autouse=True)
def isolated_export_database():
    db = Mock()
    db.query.return_value.order_by.return_value.all.return_value = []
    previous_override = app.dependency_overrides.get(get_db)
    app.dependency_overrides[get_db] = lambda: db

    yield db

    if previous_override is None:
        app.dependency_overrides.pop(get_db, None)
    else:
        app.dependency_overrides[get_db] = previous_override


def test_export_csv(
    client
):

    response = client.get(
        "/export/csv"
    )

    assert response.status_code == 200

    assert response.headers[
        "content-type"
    ].startswith(
        "text/csv"
    )

    assert (
        "attachment"
        in response.headers[
            "content-disposition"
        ]
    )

    content = (
        response.content.decode(
            "utf-8"
        )
    )

    assert (
        "id"
        in content
    )

    assert (
        "name"
        in content
    )

    assert (
        "candidate_level"
        in content
    )

    assert (
        "skill_score"
        in content
    )


def _candidate(**overrides):
    values = {
        "id": 7,
        "name": "Taylor Morgan",
        "candidate_level": "Senior",
        "skill_score": 78,
        "rule_score": 74,
        "ai_score": 92,
        "candidate_stage": "INTERVIEW",
        "ai_status": "success",
        "summary": "Evidence-supported candidate summary " * 8,
        "created_at": datetime(
            2026, 8, 14, 10, 30, tzinfo=timezone.utc
        ),
        "score_breakdown": {
            "score_version": "profile_v2",
            "professional_experience": 22,
            "achievements": 16,
            "competencies": 20,
            "certifications": 10,
            "education": 5,
            "leadership": 8,
            "evidence_quality": 5,
        },
        "resume_storage_key": "private/resumes/should-not-export.pdf",
        "resume_sha256": "sensitive-digest-should-not-export",
    }
    values.update(overrides)
    return SimpleNamespace(**values)


def test_export_excel_endpoint_returns_valid_attachment(client):
    response = client.get("/export/xlsx")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(EXCEL_CONTENT_TYPE)
    assert response.headers["content-disposition"].startswith(
        'attachment; filename="ATS_Candidates_'
    )
    assert response.headers["content-disposition"].endswith('.xlsx"')
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == [
        "Candidates",
        "Score Breakdown",
        "Legacy Scores",
    ]


def test_workbook_has_deterministic_columns_and_versioned_scores():
    profile = _candidate()
    legacy = _candidate(
        id=8,
        name="Legacy Engineer",
        skill_score=61,
        rule_score=58,
        score_breakdown={
            "python": 10,
            "sql": 8,
            "backend": 7,
            "engineering_signal": 3,
        },
    )
    workbook = load_workbook(
        build_candidates_workbook([profile, legacy])
    )

    candidates = workbook["Candidates"]
    assert tuple(
        cell.value for cell in candidates[6]
    ) == CANDIDATE_HEADERS
    assert candidates["E7"].value == "profile_v2"
    assert candidates["F7"].value == 78
    assert candidates["G7"].value is None
    assert candidates["H7"].value == 74
    assert candidates["E8"].value == "technical_v1 (legacy)"
    assert candidates["F8"].value is None
    assert candidates["G8"].value == 61
    assert candidates["I8"].value == 58
    assert candidates["F7"].data_type == "n"
    assert candidates["J7"].data_type == "n"

    profile_sheet = workbook["Score Breakdown"]
    assert tuple(
        cell.value for cell in profile_sheet[6]
    ) == PROFILE_BREAKDOWN_HEADERS
    assert profile_sheet.max_row == 7
    assert profile_sheet["A7"].value == profile.id

    legacy_sheet = workbook["Legacy Scores"]
    assert tuple(
        cell.value for cell in legacy_sheet[6]
    ) == LEGACY_BREAKDOWN_HEADERS
    assert legacy_sheet.max_row == 7
    assert legacy_sheet["A7"].value == legacy.id


def test_workbook_handles_nulls_long_text_and_private_fields_safely():
    candidate = _candidate(
        name="=HYPERLINK(\"https://example.invalid\")",
        summary=("Long summary\x00 " * 3000),
        created_at=None,
    )
    workbook = load_workbook(build_candidates_workbook([candidate]))
    sheet = workbook["Candidates"]

    assert sheet["C7"].value.startswith("'=")
    assert "\x00" not in sheet["M7"].value
    assert len(sheet["M7"].value) == 32767
    assert sheet["M7"].alignment.wrap_text is True
    assert sheet["N7"].value is None

    all_text = " ".join(
        str(cell.value or "")
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
    )
    assert "private/resumes" not in all_text
    assert "sensitive-digest" not in all_text
    assert "hashed_password" not in all_text
    assert "token_version" not in all_text
    assert "profile_image_key" not in all_text


def test_workbook_applies_professional_usability_formatting():
    workbook = load_workbook(build_candidates_workbook([_candidate()]))

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        assert sheet.freeze_panes == "A7"
        assert sheet.auto_filter.ref.startswith("A6:")
        assert sheet["A1"].value == "ATS Resume Intelligence"
        assert sheet["A6"].font.bold is True
        assert sheet["A6"].fill.fill_type == "solid"
        assert sheet.column_dimensions["A"].width > 0
        assert sheet.sheet_view.showGridLines is False

    assert len(workbook["Candidates"].conditional_formatting) > 0
